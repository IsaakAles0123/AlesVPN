"""Заполняет alesvpn-wireguard-учёт.xlsx на рабочем столе OneDrive."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(r"C:\Users\isaak\OneDrive\Рабочий стол\alesvpn-wireguard-учёт.xlsx")

HEADERS = [
    "Дата первого подключения (факт)",
    "Клиент / устройство",
    "Контакт",
    "IP в VPN",
    "Приватный ключ клиента",
    "Публичный ключ клиента",
    "Peer на сервере",
    "Заметки",
]

ROWS = [
    [
        "2026-04-15",
        "iPhone / myphone-16proMax (client-01, 10.8.0.2)",
        "",
        "10.8.0.2/32",
        "SKmPOivri8giflnX1vWp27ds+9+56xI5Dh5pJU1k12E=",
        "nZs3C99LKqlg6dA3tNdn3eUXgG0EnTmNjeDU+T+CZSI=",
        "да",
        "авторизован, handshake с сервера",
    ],
    [
        "2026-04-15",
        "client-02",
        "",
        "10.8.0.3/32",
        "ICg2HQCFmnF8vBvyNuq+W3NUK7G33ri4h4nVNePMQkM=",
        "HQYY5ZltkqTXjBo3eUC/NFFdARB+7xM/hzrrhUbID00=",
        "да",
        "авторизован, handshake с сервера",
    ],
]

COL_WIDTHS = [22, 38, 18, 14, 52, 52, 18, 36]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты WireGuard"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4B286D")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS[col - 1]

    for r, row in enumerate(ROWS, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = wrap

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(ROWS) + 1}"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"OK: {OUT}")


if __name__ == "__main__":
    main()
